import pandas as pd
import glob
import os
import logging
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import PatternFill

# Constants
OUTPUT_LOCATION = "outputs/"  # Folder containing part files
FILE_PATTERN = "*_part*.csv"  # Pattern to match all part files

# Logging Configuration
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

def clean_response(text):
    """Cleans the 'Response' text by removing unnecessary bullet points and formatting."""
    if not isinstance(text, str):
        return text  # Return as-is if not a string
    
    # Remove extra spaces between bullet points
    text = re.sub(r"\n\s*[*\-•→]+", "\n*", text)
    
    # Remove unwanted bullet points
    lines = text.split("\n")
    cleaned_lines = [
        line.strip() for line in lines  # Remove leading/trailing whitespace
        if line.strip() and  # Skip empty lines
        not line.startswith("BrewChat") and
        "XML" not in line and
        not line.startswith("Dates are answered in the format")
    ]
    
    # Replace "â€¢" with "*"
    cleaned_text = "\n".join(["\t" + line for line in cleaned_lines]).replace("â€¢", "*")  # Add tab indentation to each line
    
    return cleaned_text.strip()  # Ensure no leading/trailing whitespace

from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import PatternFill

def format_excel(file_path):
    """Applies formatting to the final Excel file, including conditional formatting for the 'Response' column."""
    wb = load_workbook(file_path)
    ws = wb.active

    # Freeze the top row
    ws.freeze_panes = "A2"

    # Auto adjust column widths for all except fixed-width columns
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        if col_letter not in ["C", "D"]:  # Skipping "EntryOriginal" and "Response"
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

    # Set fixed widths for specific columns
    ws.column_dimensions["C"].width = 30  # EntryOriginal column
    ws.column_dimensions["D"].width = 100  # Response column (adjusted to fit description + response content)

    # Wrap text in Response column
    for cell in ws["D"]:
        cell.alignment = Alignment(wrap_text=True)

    # Conditional Formatting for "Response" column
    response_col = "D2:D" + str(ws.max_row)  # Target all cells in column E, starting from row 2

    # Define color fills
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light Red
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Light Yellow
    #green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light Green

    # Apply conditional formatting based on words
    ws.conditional_formatting.add(response_col, 
        FormulaRule(formula=[f'ISNUMBER(SEARCH("inconclusive", D2))'], stopIfTrue=True, fill=red_fill))
    ws.conditional_formatting.add(response_col, 
        FormulaRule(formula=[f'ISNUMBER(SEARCH("handwritten", D2))'], stopIfTrue=True, fill=yellow_fill))
    #ws.conditional_formatting.add(response_col, 
        #FormulaRule(formula=[f'ISNUMBER(SEARCH("success", E2))'], stopIfTrue=True, fill=green_fill))

    # Format as table
    table = Table(displayName="CourtBookData", ref=ws.dimensions)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)

    # Hide gridlines
    ws.sheet_view.showGridLines = False

    wb.save(file_path)


def concatenate_parts():
    """Concatenates all part CSV files into a single final Excel file per court book and recombines entries with the same Unique ID."""
    part_files = glob.glob(os.path.join(OUTPUT_LOCATION, FILE_PATTERN))
    
    if not part_files:
        logging.info("No part files found for concatenation.")
        return
    
    # Group part files by court book ID (extract ID before '_part')
    courtbook_groups = {}
    for file in part_files:
        base_name = os.path.basename(file)
        courtbook_id = base_name.split("_part")[0]  # Extract courtbook ID
        courtbook_groups.setdefault(courtbook_id, []).append(file)
    
    # Process each court book
    for courtbook_id, files in courtbook_groups.items():
        files.sort()  # Ensure proper order (part1, part2, etc.)
        logging.info(f"Merging {len(files)} parts for court book {courtbook_id}")
        
        # Concatenate CSV files
        df_list = [pd.read_csv(f) for f in files]
        final_df = pd.concat(df_list, ignore_index=True)
        
        # Clean 'Response' column
        if "Response" in final_df.columns:
            final_df["Response"] = final_df.apply(lambda row: f"{row['EntryDescription']}\n{clean_response(row['Response'])}", axis=1)
        
        # Remove 'Part', 'EntryDescription', and 'UniqueID' columns
        final_df.drop(columns=["Part", "EntryDescription", "UniqueID"], inplace=True, errors='ignore')
        
        # Recombine entries with the same Unique ID
        if "UniqueID" in final_df.columns:
            final_df = final_df.groupby("UniqueID", as_index=False).agg({
                "EntryDate": "first",
                "Source Doc": "first",
                "EntryOriginal": "first",
                "Response": lambda x: "\n" + "\n".join(x.dropna().str.strip()).lstrip("\n") if not x.empty else "",
                "Handwritten": "first",
                "TimeProcessed": "first"
            })
        
        # Drop 'UniqueID' before saving
        final_df.drop(columns=["UniqueID"], inplace=True, errors='ignore')

        # Define final column order (without UniqueID)
        final_df = final_df[["Source Doc", "EntryDate", "EntryOriginal", "Response", "Handwritten", "TimeProcessed","LineID"]]
        
        # Save as Excel for better formatting
        final_filename = os.path.join(OUTPUT_LOCATION, f"{courtbook_id}_chronology.xlsx")
        final_df.to_excel(final_filename, index=False, engine="openpyxl")
        
        # Apply Excel formatting
        format_excel(final_filename)
        
        # Log summary
        logging.info(f"Final file saved: {final_filename} ({len(final_df)} unique entries processed)")
 
if __name__ == "__main__":
    concatenate_parts()